import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os

#Loading Excel sheet
excelPath = os.path.join(r"C:\Users\Deepak Pandya\Documents\Learning\Python\Projects\attendance.xlsx")
book = openpyxl.load_workbook(excelPath)
print(book.sheetnames)
#select sheet
sheet = book['AttendanceSheet']

r = sheet.max_row

resp =1
c= sheet.max_column

#List of studnet to remind who has taken 1 leaves
l1=[]

#list of students to remind with warning 
l2=[]

#list of defaulters
l3=[]

#staff mail id
staff_email=["dpandya59@gmail.com"]

#Warning Message
m1="WARNING!! You can take only more day leave for CI Class"
m2="WARNING!! You can take only more day leave for Python Class"
m3="WARNING!! You can take only more day leave for DM Class"

def savefile():
    book.save(excelPath)
    print("Saved!!")

def check(no_of_days,row_num,b):
    global staff_email
    global l2
    global l3

    for student in range(0,len(row_num)):
        if no_of_days[student]==2:
            l1.append(sheet.cell(row=row_num[student],column=2).value)
            if b == 1:
                a=m1
            elif b == 2:
                b=m2
            else:
                c=m3
        elif no_of_days[student]>2:
            if b == 1:
                l2= l2+str(sheet.cell(row=row_num[student],column=1).value)
                l3.append(sheet.cell(row=row_num[student],column=2).value)
                subject  ="CI"
            elif b == 2:
                l2=l2+str(sheet.cell(row=row_num[student],column=1).value)
                l3.append(sheet.cell(row=row_num[student],column=2).value)
                subject ="Python"
            else:
                l2=l2+str(sheet.cell(row=row_num[student],column=2).value)
                l3.append(sheet.cell(row=row_num[student],column=3).value)
                subject = "Data Mining"
        if l2!="" and len(l3)!=0:
            msg1="You have lack of attendance in "+subject+" !!!"
            msg2="Following students has lack of attendance in your subject "+l2
            mailstu(l3,msg1)
            staff_id =staff_email[b-1]
            mailstaff(staff_id,msg2)

def mailstu(li,msg):
    print("Attendance Report")
    for i in range (0,len(li)):
        print(li[i]+" - "+msg)

def mailstaff(mail_id,msg):
    print("Lack of attendance report:")
    print(mail_id+" - "+msg)

while resp == 1:
    print("1--->CI\n2--->python\n3--->DM")
    y=int(input("Enter subject:"))
    no_of_absentees=int(input("No. of absentees:"))

    if(no_of_absentees>1):
        x=list(map(int,(input('Roll nos:').split(' '))))
    else:
        x=[int(input('roll no:'))]

    row_num=[]
    no_of_days=[]

    for student in x:
        print("y, r: ",y,r)
        for i in range(2,r+1):
            print("i: ",i)
            if y == 1:
                if sheet.cell(row=i,column=1).value is student:
                    m = sheet.cell(row=i,column=3).value
                    m=m+1
                    sheet.cell(row=i,column=3).value=m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
            elif y == 2:
                if sheet.cell(row=i,column=1).value is student:
                    m=sheet.cell(row=i,column=4).value
                    m = m+1
                    sheet.cell(row=i,column=4).value =m+1
                    no_of_days.append(m)
                    row_num.append(i)
            elif y == 3:
                if sheet.cell(row=i,column=1).value is student:
                    m=sheet.cell(row=i,column=5).value
                    m=m+1
                    sheet.cell(row=i,column=5).value = m+1
                    no_of_days.append(m)
                    row_num.append(i)
        check(no_of_days,row_num,y)
        resp=int(input("another subject? 1--> Yes, 0--> No"))